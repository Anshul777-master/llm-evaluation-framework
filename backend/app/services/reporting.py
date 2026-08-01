import csv
import html
import io
import json
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..models import Evaluation


def evaluation_payload(evaluation: Evaluation) -> dict[str, Any]:
    return {
        "id": evaluation.id,
        "name": evaluation.name,
        "model": evaluation.model_slug,
        "dataset": evaluation.dataset_name,
        "trust_score": evaluation.trust_score,
        "grade": evaluation.grade,
        "risk_level": evaluation.risk_level,
        "scores": json.loads(evaluation.scores_json),
        "recommendation": evaluation.recommendation,
        "results": json.loads(evaluation.results_json),
        "created_at": evaluation.created_at.isoformat(),
    }


def csv_report(evaluation: Evaluation) -> bytes:
    payload = evaluation_payload(evaluation)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Sentinel AI evaluation report"])
    writer.writerow(["Evaluation", payload["name"]])
    writer.writerow(["Model", payload["model"]])
    writer.writerow(["Trust score", payload["trust_score"]])
    writer.writerow(["Grade", payload["grade"]])
    writer.writerow(["Risk", payload["risk_level"]])
    writer.writerow([])
    writer.writerow(["Dimension", "Score"])
    for name, score in payload["scores"].items():
        writer.writerow([name, score])
    writer.writerow([])
    writer.writerow(["Prompt", "Response", "Latency (ms)", "Tokens", "Flags"])
    for result in payload["results"]:
        writer.writerow([
            result["prompt"],
            result["response"],
            result["execution_time_ms"],
            result["token_usage"],
            "; ".join(flag["category"] for flag in result["flags"]),
        ])
    return output.getvalue().encode("utf-8-sig")


def xlsx_report(evaluation: Evaluation) -> bytes:
    payload = evaluation_payload(evaluation)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Sentinel AI evaluation report"])
    summary.append(["Evaluation", payload["name"]])
    summary.append(["Model", payload["model"]])
    summary.append(["Trust score", payload["trust_score"]])
    summary.append(["Grade", payload["grade"]])
    summary.append(["Risk", payload["risk_level"]])
    summary.append([])
    summary.append(["Dimension", "Score"])
    for name, score in payload["scores"].items():
        summary.append([name.title(), score])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 60

    raw = workbook.create_sheet("Raw outputs")
    raw.append(["Prompt", "Response", "Latency (ms)", "Tokens", "Flags"])
    for result in payload["results"]:
        raw.append([
            result["prompt"],
            result["response"],
            result["execution_time_ms"],
            result["token_usage"],
            ", ".join(flag["category"] for flag in result["flags"]),
        ])
    raw.column_dimensions["A"].width = 48
    raw.column_dimensions["B"].width = 90
    for row in raw.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def html_report(evaluation: Evaluation) -> bytes:
    payload = evaluation_payload(evaluation)
    dimensions = "".join(
        f"<div class='metric'><span>{html.escape(name.title())}</span><strong>{score:.1f}</strong></div>"
        for name, score in payload["scores"].items()
    )
    rows = "".join(
        f"<tr><td>{html.escape(result['prompt'])}</td><td>{html.escape(result['response'])}</td><td>{result['execution_time_ms']} ms</td></tr>"
        for result in payload["results"]
    )
    document = f"""<!doctype html><html><head><meta charset='utf-8'><title>{html.escape(payload['name'])}</title><style>
    body{{font-family:Arial,sans-serif;color:#121619;max-width:1000px;margin:40px auto;padding:0 24px}}h1{{font-family:Georgia,serif;font-size:42px}}.hero{{background:#f6f3ec;border:1px solid #d8d7cf;padding:24px;border-radius:14px}}.score{{font-size:64px;color:#118a68}}.metrics{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:24px 0}}.metric{{padding:14px;border:1px solid #ddd;border-radius:10px}}.metric span{{display:block;color:#66706a;font-size:12px}}.metric strong{{font-size:28px}}table{{width:100%;border-collapse:collapse}}th,td{{padding:12px;text-align:left;border-bottom:1px solid #ddd;vertical-align:top}}th{{background:#2457d6;color:white}}p{{line-height:1.6}}
    </style></head><body><p>Sentinel AI · Model assurance report</p><div class='hero'><h1>{html.escape(payload['name'])}</h1><div class='score'>{payload['trust_score']:.1f} · Grade {payload['grade']}</div><p>{html.escape(payload['recommendation'])}</p></div><div class='metrics'>{dimensions}</div><h2>Raw outputs</h2><table><thead><tr><th>Prompt</th><th>Response</th><th>Latency</th></tr></thead><tbody>{rows}</tbody></table></body></html>"""
    return document.encode("utf-8")


def pdf_report(evaluation: Evaluation) -> bytes:
    payload = evaluation_payload(evaluation)
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Sentinel AI · Model assurance report", styles["Overline"]),
        Spacer(1, 5 * mm),
        Paragraph(payload["name"], styles["Title"]),
        Paragraph(f"Model: {payload['model']} · Dataset: {payload['dataset']}", styles["BodyText"]),
        Spacer(1, 5 * mm),
        Paragraph(f"Trust score: {payload['trust_score']:.1f} · Grade {payload['grade']} · {payload['risk_level']} risk", styles["Heading2"]),
        Paragraph(payload["recommendation"], styles["BodyText"]),
        Spacer(1, 7 * mm),
    ]
    score_table = Table([["Dimension", "Score"]] + [[name.title(), f"{score:.1f}"] for name, score in payload["scores"].items()], colWidths=[105 * mm, 45 * mm])
    score_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2457D6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8D7CF")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.extend([score_table, Spacer(1, 8 * mm), Paragraph("Sample outputs", styles["Heading2"])])
    for result in payload["results"][:8]:
        story.append(Paragraph(f"<b>Prompt:</b> {html.escape(result['prompt'])}", styles["BodyText"]))
        story.append(Paragraph(f"<b>Response:</b> {html.escape(result['response'])}", styles["BodyText"]))
        story.append(Spacer(1, 4 * mm))
    document.build(story)
    return output.getvalue()


REPORTERS = {"csv": csv_report, "xlsx": xlsx_report, "html": html_report, "pdf": pdf_report}


def generate_report(evaluation: Evaluation, file_format: str) -> tuple[bytes, str, str]:
    if file_format not in REPORTERS:
        raise ValueError("Supported formats: pdf, xlsx, csv, html")
    data = REPORTERS[file_format](evaluation)
    media_types = {
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
        "html": "text/html",
    }
    safe_name = "".join(character if character.isalnum() or character in "-_" else "-" for character in evaluation.name.lower()).strip("-")
    return data, media_types[file_format], f"{safe_name or 'evaluation-report'}.{file_format}"
